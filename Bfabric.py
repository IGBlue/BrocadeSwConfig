#
#   Program: Bfabric.py
#
#   Author: Ian Gray
#   Contact: iangray100@gmail.com
#
#   This program is designed to generate Brocade fibre channel switch commands to configure zoning. Input is taken
#   from a Excel spreadsheet file containing interface information or the equivalent CSV file.
#     Col 1. Primary node name
#     Col 2. Primary interface name
#     Col 3. Secondary interface name - this is optional
#     Col 4. The fabric id (A or B). Dual fabric configuration is assumed
#     Col 5. Initiator or Target indicator (use I or T)
#     Col 6. World Wide Port Name - 8 x hex characters pairs separated by a colon
#     Col 7. World Wide Node Name- as for WWPN. This is not used. Documentation only
#
#   Input: A valid XLSX or CSV file normally exported from Excel or similar.
#       File location and name will be prompted for
#
#   Output: Six files containing alias, zone and configuration commands. One set of files per fabric.
#
#   Versions:
#     1   19/04/2020  Base version - based on Fabric.py and Fabricxl.py this variation
#                     takes either a .CSV file or an .XLSX file as input and saves having to
#                     support two programs. The input file type infers the source file and 
#                     subsequent processing.
#
#     2   27/06/2020  Added optional fabric/cable check and duplicate names.
#

import os  # Allows for file validation
from pathlib import Path  # Allows parsing of filename
# import pdb                        # Trace routines - pdb.set_trace()
from datetime import datetime  # Date/time retrieval modules
from openpyxl import load_workbook  # Excel workbook functions
import csv  # Allow CSV files to be read


#
# Define functions first -----------------------------------------------------------
#

#
#  Strict checking only allows Brocade format WWIDs i.e 8 pairs of hex characters separated by a colon
#  Returns None if data is invalid or the WWID if it is valid
#


def strictwwid(wwid):
    """Validity checks WWID for invalid format using strict rules - must be Brocade format"""
    if len(wwid) != 23:  # WWPN must be 23 characters long
        print("WWID has invalid length " + wwid)
        return None
    # Colon separators must be in the correct place
    if wwid[2:3] != ":" and wwid[5:6] != ":" and wwid[8:9] != ":" and wwid[11:12] != ":" \
            and wwid[14:15] != ":" and wwid[17:18] != ":" and wwid[20:21] != ":":
        print(("WWID invalid format - colon not where expected " + wwid))
        return None
    # Remove colons from expected locations in wwid string.
    hexwid = wwid[0:2] + wwid[3:5] + wwid[6:8] + wwid[9:11] + wwid[12:14] + wwid[15:17] + wwid[18:20] + wwid[21:23]
    # Only hex characters allowed in wwpn after extracting colons
    for nibble in hexwid:
        if (nibble < "0" or nibble > "f") or (nibble > "9" and nibble < "a"):
            print("WWID has invalid hex character " + wwid)
            return None

    return wwid  # Looks good so return WWID


#
#  Loosewwid takes a WWID string and extracts hex characters. If 16 hex characters are found then
#  WWID is taken as OK and it is returned in Brocade format. Any non hex characters are ignored and
#  assumed to be fromatting characters from the original data source.
#  Returns None if data is invalid or the WWID if it is valid
#


def loosewwid(wwid):
    """Validity checks WWID for invalid format using loose rules - any hex values used:"""

    hexstring = ""

    for userchar in wwid:  # Get each character in the input and isolate hex characters
        userchar = userchar.lower()  # For consistency convert to lower case
        # Only hex values 0-9 and a-f allowed
        if (userchar >= "0" and userchar <= "9") or (userchar >= "a" and userchar <= "f"):
            hexstring = hexstring + userchar  # Form an unformatted string
    #
    # Should now have a string of length 16. If we haven't then user input was invalid
    #

    if len(hexstring) != 16:
        print("WWID has invalid length " + wwid)
        return None
    else:
        #
        # Brocade format 2 chars with colon separator
        #
        brocadefmt = hexstring[0:2] + ":" \
                     + hexstring[2:4] + ":" \
                     + hexstring[4:6] + ":" \
                     + hexstring[6:8] + ":" \
                     + hexstring[8:10] + ":" \
                     + hexstring[10:12] + ":" \
                     + hexstring[12:14] + ":" \
                     + hexstring[14:16]
        return brocadefmt  # Return accepted WWID in Brocade format


#
# We need to combine primary node name, primary interface name, optionally a secondary interface name
# and the fabric identifier into a string to be used as the alias name.
#


def alias_format(linelist):
    """Takes an indicated list entry from all_fab and returns a string to be used as the alias name"""
    ali = "ali_" + linelist[0]  # Primary node identifier
    if linelist[1] != "":  # Primary interface id
        ali = ali + "_" + linelist[1]
    if linelist[2] != "":  # Secondary interface id
        ali = ali + "_" + linelist[2]
    ali = ali + "_F" + linelist[3]  # Append "F" for Fabric + fabric id
    return ali  # Return alias name


#
#
# End of functions -----------------------------------------------------------------
#

#
# Main processing begins
#
# Section 1 - read all records from the XLSX file and validity check the data (as best we can).
# For each valid record put the contents into a list (all_fab).
# If any validation errors are found report them and halt the program.
#
# Get the file information and working directory
#
#
while True:
    folder = input("Enter the working directory: ")  # Directory for XLSX file and output files
    if folder == "":  # Exit if no input
        exit(1)
    if os.path.isdir(folder):  # Check only directory name specified
        break
    else:
        print("Folder/directory not valid or filename specified - " + folder)
        print("")

while True:
    csvfile = input("CSV or XLSX filename in " + folder + ": ")  # Data filename request
    if csvfile == "":  # Exit if no input
        exit(1)
    infile = folder + "\\" + csvfile  # Input file is directory + filename

    filetype = Path(csvfile).suffix  # Check we only have CSV or XLSX file
    filetype = filetype.lower()
    if filetype == ".csv" or filetype == ".xlsx":
        pass
    else:
        print("Incorrect file type - " + csvfile)
        print("")
        continue

    if os.path.isfile(infile):  # Check file exists
        break
    else:
        print("Invalid filename or does not exist - " + infile)
        print("")

#
# Form the strings used to create the output files
#
afabali_file = folder + "\\" + "Afab_ali.txt"  # Fabic A alias definitions
bfabali_file = folder + "\\" + "Bfab_ali.txt"  # Fabric B alias definitions

afabzon_file = folder + "\\" + "Afab_zon.txt"  # Fabric A zone definitions
bfabzon_file = folder + "\\" + "Bfab_zon.txt"  # Fabric B zone definitions

afabcfg_file = folder + "\\" + "Afab_cfg.txt"  # Fabric A configuration definition
bfabcfg_file = folder + "\\" + "Bfab_cfg.txt"  # Fabric B configuration definition

#
# Delete the files we are about to generate (if present) to ensure that old versions are not used
#
if os.path.exists(afabali_file):
    os.remove(afabali_file)
if os.path.exists(bfabali_file):
    os.remove(bfabali_file)
if os.path.exists(afabzon_file):
    os.remove(afabzon_file)
if os.path.exists(bfabzon_file):
    os.remove(bfabzon_file)
if os.path.exists(afabcfg_file):
    os.remove(afabcfg_file)
if os.path.exists(bfabcfg_file):
    os.remove(bfabcfg_file)

strict = ""  # Used to flag Strict or Loose validity checking

while strict == "":  # Get validation option
    strict = input("Validity check Strict, Loose or Help (S,L,?) ").upper()

    if strict == "?":  # Tell user the differences
        print("Strict checking will only accept WWIDs in Brocade format\n")
        print("Loose will accept any string but will convert embedded hex characters into Brocade format")
        print("Any non hex characters are ignored\n")
        print("There must be 16 hex characters in either string")
        strict = ""
    elif strict == "S" or strict == "L":
        continue
    else:
        strict = ""  # User can't type
#
# Ask user if fabric/cable checks should be done
#
cabchk = ""
while cabchk == "":
    cabchk = input("Do you want to flag possible cable fabric/issues: (Y,N,?) ").upper()  # Ask 

    if cabchk == "?":  # Exit if no input
        print("The input rows will be sorted on the three name columns A,B,C")
        print("and the WWPN col F. If the naming is consistent and the")
        print("manufacturer uses sequential WWPNs for each device then the")
        print("Fabric indicator (col D)  should alternate between A and B. If it")
        print("does not then a possible cabling or naming problem is indicated.")
        cabchk = ""
    elif cabchk == "Y" or cabchk == "N":
        continue
    else:
        print("Unrecognised input")
        print("")
        cabchk = ""

#
# In this section we are reading all non blank lines and constructing an internal list
# of these items - note this all_fab list is a list of lists.
# Remember that subscripts in Python are relative to zero
#
header_passed = False  # Flag value to bypass 1st header row
errors_found = False  # So we don't process output if invalid data found
all_fab = []  # Create an empty list to contain read data
wwidct = 0  # To count the number of wwid entries

if filetype == ".csv":
    with open(infile) as f:  # Open input file
        for rec in csv.reader(f):  # Read and process all recs in CSV file
            node = rec[0].strip()
            primaryif = rec[1].strip()  # Strip removes leading/trailng spaces
            subif = rec[2].strip()
            fabric = rec[3].upper()  # Make fabric id consistent
            fabric = fabric.strip()
            initgt = rec[4].upper()  # Make sure case is consistent
            initgt = initgt.strip()
            wwpn = rec[5].lower()  # Force lower case for hex characters
            wwpn = wwpn.strip()

            if not header_passed:  # Ignore first line as column headers from spreadsheet
                header_passed = True
                continue
            if len(node) == 0:  # If no node name whole line assumed empty and ignored
                continue
            all_fab.append([node, primaryif, subif, fabric, initgt, wwpn])
else:
    xlworkbook = load_workbook(infile)  # Load the spreadsheet
    xlsheets = xlworkbook.active  # Get sheet names in case of multiples
    xlsheet = xlworkbook.active  # Set the active sheet as the one to work on
    xlmaxrow = xlsheet.max_row  # Get the number of rows in the sheet
    # Read and process all recs in spreadsheet active sheet
    for rec in xlsheet.iter_rows(min_row=2,  # Min row = 2 - row 1 assumed to be header
                                 max_row=xlmaxrow,
                                 min_col=1,
                                 max_col=6,
                                 values_only=True):
        if rec[0] is None or rec[0] == "":  # openpyxl returns None for no value in cell
            continue  # Assume blank line in this case
        node = rec[0].strip()
        primaryif = rec[1].strip()  # Strip removes leading/trailng spaces
        if rec[2] is None:  # Subif is the only column which can have no value
            subif = ""  # None values cause issues with sort functions later
        else:
            subif = rec[2].strip()
        fabric = rec[3].upper()  # Make fabric id consistent
        fabric = fabric.strip()
        initgt = rec[4].upper()  # Make sure case is consistent
        initgt = initgt.strip()
        wwpn = rec[5].lower()  # Force lower case for hex characters
        wwpn = wwpn.strip()
        all_fab.append([node, primaryif, subif, fabric, initgt, wwpn])
if filetype == ".xlsx":
    del xlworkbook  # Finished with workbook
#
# The all_fab list contains one entry for each source row but the contents have not
# been validity checked. This next section reads all_fab and checks those items it can
#        

for rec in all_fab:
    node = rec[0]
    primaryif = rec[1]
    subif = rec[2]
    fabric = rec[3]
    initgt = rec[4]
    wwpn = rec[5]

    if strict == "S":  # Check wwpn has valid format
        wwidok = strictwwid(wwpn)
        if wwidok is None:
            errors_found = True
            continue
    elif strict != "S":  # Loose checking
        wwidok = loosewwid(wwpn)
        if wwidok is None:
            errors_found = True
            continue
        else:
            wwpn = wwidok

    # Must be Initiator or Target
    if initgt != "I" and initgt != "T":
        print("Invalid Initiator/Target value - must be I or T " + wwpn)
        errors_found = True
        continue

    # Fabric must be A or B
    if fabric != "A" and fabric != "B":
        print("Fabric identifier must be A or B " + wwpn)
        errors_found = True
        continue
    # Add details to all_fab list

    wwidct += 1  # Update the valid record count

#
# Section 2a
# Sort the all_fab list into wwpn order so we can check for duplicate wwpns
#
sort_WWPN = lambda swwpn: swwpn[5]  # Sort funcion to return key value wwpn
all_fab.sort(key=sort_WWPN)  # Sort list on wwpn

# Now process the list checking for duplicate wwpn entries
last_wwpn = ""  # Last wwpn value
for devrow in all_fab:  # Process each row in newly sorted list
    dr_wwpn = devrow[5]  # Extract wwpn from sorted list

    if dr_wwpn != last_wwpn:  # If we already found it then a duplicate exists
        last_wwpn = dr_wwpn  # Else update the wwpn data
    else:
        print("Duplicate wwpn found " + dr_wwpn)
        errors_found = True
#
# Section 2b
# Sort the all_fab list on Node,i/f,subif,wwpn to check for cabling inconsistencies.
# In addition (and since we are in the correct order) check for duplicate names.
# In order to be consistent cols A,B,C have a length set to 10 characters. The WWPN should have
# a consistent length.
# Define the sort key function first.
#
sort_cabkeys = lambda cabval: "{:<10}".format(cabval[0]) + \
                              "{:<10}".format(cabval[1]) + \
                              "{:<10}".format(cabval[2]) + \
                              cabval[5]
all_fab.sort(key=sort_cabkeys)  # Sort the data for fabric/cable/name reporting.
#
# Data is now sorted on name columns + WWPN. First check that there are no duplicate names
#

prv_node = " "  # Set initial previous line details
prv_iface = " "
prv_subif = " "
dupname_error = False

for devrow in all_fab:  # Check current row values against previous row
    if devrow[0] == prv_node and \
            devrow[1] == prv_iface and \
            devrow[2] == prv_subif:
        print("Duplicate name found: " + devrow[0] +
              " " + devrow[1] +
              " " + devrow[2])
        print("")
        errors_found = True  # Mark error found
        dupnam_error = True  # Need to bypass cable check
    prv_node = devrow[0]
    prv_iface = devrow[1]
    prv_subif = devrow[2]
#
# Go through the sorted list again (if requested) and check that the fabric indicators alternate
# Note - if duplicate name check finds errors it causes confusion in the cable check below
# so don't do this if errors found.
#
if cabchk == "Y" and not dupname_error:  # Only if user has elected to do this
    fabriccmp = " "  # Set initial values
    last_devrow = ""
    for devrow in all_fab:  # Process all rows
        fabric = devrow[3]  # Extract the fabric indicator
        if fabric == fabriccmp:  # Is it the same as previous list item
            print("Possible cable misconfiguration detected")  # Yes - so report it
            print(last_devrow)  # Print this and the previous list item
            print(devrow)
            print("")
            errors_found = True
        last_devrow = devrow
        fabriccmp = fabric

#
# At this point we should either have found a data inconsistency (in which case stop now)
# or we think all the data is good (in which case carry on).
#

if errors_found:
    print("Data errors found - output files NOT produced")
    print("Note errors and hit return to terminate program")
    print("")
    rtn = input("")
    print("Exiting")
    exit(1)
#
# Section 3.
# Sort the all_fab list by node and interface ids. We do this because we need to create
# an alias and it's just nice to have them in alphabetical order.
#
sort_ALIif = lambda salitgt: salitgt[0] + salitgt[1] + salitgt[2]
all_fab.sort(key=sort_ALIif)  # Sort list on IT + interface info
#
# Open output files for alias create commands
#
afabali = open(afabali_file, "w")  # These files hold alias commands
afabali.write("# Alias create commands for fabric A\n")
bfabali = open(bfabali_file, "w")  # One file for each fabric
bfabali.write("# Alias create commands for fabric B\n")

aliact = 0  # Counts for number of alias ecords generated
alibct = 0

for devrow in all_fab:  # For each valid WWID
    aliname = alias_format(devrow)  # Create alias name from info in list entry
    # Now format an aliCreate line as per Brocade CLI
    aliline = 'aliCreate ' + '"' + aliname + '", "' + devrow[5] + '"\n'
    if devrow[3] == "A":
        afabali.write(aliline)  # Write to fabric A file
        aliact += 1  # Increment fab A alias count
    else:
        bfabali.write(aliline)  # Write to fabric B file
        alibct += 1  # Increment fab B alias count

afabali.close()  # Finshed with alias files
bfabali.close()

# Print alias counts
print("")
print(str(aliact) + " alias records were written for fabric A")
print(str(alibct) + " alias records were written for fabric B\n")

#
# Sort the all_fab list into I/T order followed by interface ids. We do this because we need to produce
# a zone configuration record where every initiator interface is connected to target devices on the same fabric
#
sort_ITif = lambda sinitgt: sinitgt[4] + sinitgt[0] + sinitgt[1] + sinitgt[2]
all_fab.sort(key=sort_ITif)  # Sort list on Init/Tgt + interface info

# Now split the list so we have a list of initiators and a list of targets.
# Note validation only allows I or T

INI_list = []  # Create empty Initiator list
TGT_list = []  # Create empty Target list

for devrow in all_fab:  # Extract each row and assign to INI or TGT list as appropriate
    if devrow[4] == "I":  # Get Init/Tgt identifier
        INI_list.append(devrow)
    else:
        TGT_list.append(devrow)  # Assume target as we only allow I or T

#
# Section 4.
# We should now have produced the alias commands and have split the original all_fab list
# into two new lists - one for initiators and one for targets
#
# For each initiator generate a zone record where
#   a. the initiator and target are in the same fabric
#   b. the initiator and primary interface have a common target node name
#
# Note that the zones are created with a zoneCreate command but common targets are added to the
# zone with a zoneAdd command.
#
afabzon = open(afabzon_file, "w")  # These files hold zone commands
afabzon.write("# Zone create commands for fabric A\n")

bfabzon = open(bfabzon_file, "w")
bfabzon.write("# Zone create commands for fabric B\n")

afabcfg = open(afabcfg_file, "w")
afabcfg.write("# Switch config commands for fabric A\n")
afabcfg.write("cfgClear\n")
afabcfg.write("cfgDisable\n")

bfabcfg = open(bfabcfg_file, "w")
bfabcfg.write("# Switch config commands for fabric B\n")
bfabcfg.write("cfgClear\n")
bfabcfg.write("cfgDisable\n")

#
# Generate a new configuration name
#
cfgname = "cfg" + datetime.today().strftime('%Y-%m-%d')

last_ininode = None  # Remember the last initiator alias name for changes to zone command
last_iniprime = None  # Remember the last primary initiator node name
last_tgtnode = None  # Remember the last target node name for changes to zone name

cfga_count = 0  # We need to know when we write the first zoneCreate to each fabric
cfgb_count = 0  # in order to format the correct cfgCreate/cfgAdd commands
#
# For each initiator create a zone record for each target that is in the same fabric.
#
for inirow in INI_list:  # For each initiator record
    inode = inirow[0]  # Extract the node name
    iprime = inirow[1]  # The primary interface identifier
    initali = alias_format(inirow)  # Need the initiator alias for zoneCreate command
    for tgtrow in TGT_list:  # For each target record

        if inirow[3] != tgtrow[3]:  # Check that initiator and target are in the same fabric
            continue  # Ignore if different fabrics

        else:
            tnode = tgtrow[0]  # Isolate the target node name
            # Construct the zone name from init node, init if, tgt node
            zoname = 'zon_' + inode + '_' + iprime + '_' + tnode
            zonali = alias_format(tgtrow)  # Get the target alias name
            #
            # If the initiator node, initiator primary i/f or the target node have changed then we
            # need to use a zoneCreate command otherwise a zoneAdd
            #
            if inode == last_ininode and \
                    iprime == last_iniprime and \
                    tnode == last_tgtnode:  # No change = zoneAdd command
                zonline = 'zoneAdd    "' + zoname + '", "' + zonali + '"\n'
                zonline2 = None
            else:  # Tgt node name change = zoneCreate command
                zonline = 'zoneCreate "' + zoname + '", "' + initali + '"\n'
                zonline2 = 'zoneAdd    "' + zoname + '", "' + zonali + '"\n'
            #
            # In addition add entries to the fabric cfg commands but be aware the first command
            # per fabric file has to be a cfgCreate - all other additions are cfgAdd commands.
            # cfg commands are only added to when a zoneCreate command is generated so the condition
            # is when a zonline2 is written but this is the first zoneCreate for each fabric.
            #
            if inirow[3] == "A":  # If fabric A
                afabzon.write(zonline)  # Write to fabric A file
                if zonline2 is not None:
                    afabzon.write(zonline2)
                    cfga_count += 1
                    if cfga_count == 1:
                        afabcfg.write('cfgCreate "' + cfgname + '", "' + zoname + '"\n')
                    else:
                        afabcfg.write('cfgAdd    "' + cfgname + '", "' + zoname + '"\n')

            else:  # Else fabric B
                bfabzon.write(zonline)  # Write to fabric B file
                if zonline2 is not None:
                    bfabzon.write(zonline2)
                    cfgb_count += 1
                    if cfgb_count == 1:
                        bfabcfg.write('cfgCreate "' + cfgname + '", "' + zoname + '"\n')
                    else:
                        bfabcfg.write('cfgAdd    "' + cfgname + '", "' + zoname + '"\n')

            last_ininode = inode  # Update the historical values used for zone change
            last_iniprime = iprime
            last_tgtnode = tnode

afabzon.close()  # Finshed with zone files
bfabzon.close()
afabcfg.write("cfgSave")  # Finished with cfg files
bfabcfg.write("cfgSave")
afabcfg.close()
bfabcfg.close()
#
# Remind user where output files are located
#
print("The following files have been generated:")
print("")
print("Fabric A")
print(afabali_file + " - Alias commands")
print(afabzon_file + " - Zone commands")
print(afabcfg_file + " - Configure commands")

print("")
print("Fabric B")
print(bfabali_file + " - Alias commands")
print(bfabzon_file + " - Zone commands")
print(bfabcfg_file + " - Configure commands")

print("")
print(str(wwidct) + " records in total were processed")

rtn = input("Note messages - return to finish the program")
